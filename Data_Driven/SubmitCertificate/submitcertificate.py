# Generated by Selenium IDE
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import Select

from selenium.common.exceptions import NoSuchElementException


class FileExcelReader:
    file = ""
    sheetName = ""

    def __init__(self, file, sheetName):
        self.file = file
        self.sheetName = sheetName

    def getRowCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_row)

    def getColumnCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_column)

    def readData(self, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return sheet.cell(row=rownum, column=colnum).value

    def writeData(self, data, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        wordbook.save(self.file)


class SubmitCer():
    def setup_method(self):
        self.driver = webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get("https://mybk.hcmut.edu.vn/my/logoutSSO.action")
        self.driver.get(
            'https://sso.hcmut.edu.vn/cas/login?service=https://mybk.hcmut.edu.vn/my/homeSSO.action')
        username = self.driver.find_element(By.NAME, "username")
        password = self.driver.find_element(By.NAME, "password")
        submitBtn = self.driver.find_element(By.NAME, "submit")

        username.send_keys("tan.lamcs1001")
        password.send_keys("lnt@H1720")
        submitBtn.click()

    def teardown_method(self):
        self.driver.get("https://mybk.hcmut.edu.vn/my/homeSSO.action")
        logoutBtn = self.driver.find_element(
            By.XPATH, "/html/body/div[2]/div/div/div[2]/div/a[1]")
        logoutBtn.click()
        self.driver.quit()

    def test_submitcertificate(self, listening, reading, expectedResult):
        self.driver.get(
            "https://mybk.hcmut.edu.vn/apps/src/nopccav/index.aspx")
        self.driver.find_element(By.XPATH, '//*[@id="bnt_themmoi"]').click()
        time.sleep(1)
        selectType = Select(self.driver.find_element(By.ID, 'cbo_loaicc'))
        selectType.select_by_value("TOEIC_1")
        time.sleep(1)
        id = self.driver.find_element(
            By.NAME, "ctl00$ContentPlaceHolder1$txt_TOEIC_IdNumber")
        date = self.driver.find_element(
            By.NAME, "ctl00$ContentPlaceHolder1$txt_TOEIC_TestDate")
        listeningInp = self.driver.find_element(
            By.NAME, "ctl00$ContentPlaceHolder1$txt_TOEIC_Listening")
        readingInp = self.driver.find_element(
            By.NAME, "ctl00$ContentPlaceHolder1$txt_TOEIC_Reading")
        total = self.driver.find_element(
            By.NAME, "ctl00$ContentPlaceHolder1$txt_TOEIC_TotalScore")
        id.send_keys("1111111111")
        date.send_keys("23/11/2022")
        readingInp.send_keys(reading)
        listeningInp.send_keys(listening)
        total.send_keys(str(int(reading) + int(listening)))
        btn = self.driver.find_element(By.XPATH, '//*[@id="bnt_thoinhap"]')
        btn.click()
        if (expectedResult == "Success"):
            self.driver.find_element(By.XPATH, '//*[@id="lst_danhsachccnn"]')
        elif expectedResult == "Fail":
            self.driver.find_element(By.XPATH, '//*[@id="lst_danhsachccnn"]')
        return


if __name__ == "__main__":
    excel = FileExcelReader('SecB_submitcert_data.xlsx', 'Sheet1')

    test = SubmitCer()
    test.setup_method()
    nRows = excel.getRowCount()
    for row in range(2, nRows + 1):
        listening = excel.readData(row, 1)
        reading = excel.readData(row, 2)
        expectedResult = excel.readData(row, 3)

        try:
            result = test.test_submitcertificate(
                listening, reading, expectedResult)
            excel.writeData("Passed", row, 4)
        except:
            excel.writeData("Failed", row, 4)

    test.teardown_method()
