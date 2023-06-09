# Generated by Selenium IDE
import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import Select

from selenium.common.exceptions import NoSuchElementException

class FileExcelReader:
    file = ""
    sheetName = ""

    def __init__(self, file, sheetName):
        self.file = file
        self.sheetName = sheetName

    def getRowCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_row)

    def getColumnCount(self):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return (sheet.max_column)

    def readData(self, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        return sheet.cell(row=rownum, column=colnum).value

    def writeData(self, data, rownum, colnum):
        wordbook = openpyxl.load_workbook(self.file)
        sheet = wordbook[self.sheetName]
        sheet.cell(row=rownum, column=colnum).value = data
        wordbook.save(self.file)

class TestPrintTranscript():
  def setup_method(self):
    self.driver = webdriver.Chrome()
    self.driver.maximize_window()
    self.driver.get("https://mybk.hcmut.edu.vn/my/logoutSSO.action")
    self.driver.get('https://sso.hcmut.edu.vn/cas/login?service=https://mybk.hcmut.edu.vn/my/homeSSO.action')
    username = self.driver.find_element(By.NAME,"username")
    password = self.driver.find_element(By.NAME,"password")  
    submitBtn = self.driver.find_element(By.NAME,"submit")  
    
    username.send_keys("son.cuthanh27")
    password.send_keys("02072002")
    submitBtn.click()
  
  def teardown_method(self):
    self.driver.get("https://mybk.hcmut.edu.vn/my/homeSSO.action")
    logoutBtn = self.driver.find_element(By.XPATH,"/html/body/div[2]/div/div/div[2]/div/a[1]")
    logoutBtn.click()
    self.driver.quit()
  

  def test_printtranscript(self, selectCampus, quantity, confirm, expectedResult):
    self.driver.get('https://mybk.hcmut.edu.vn/apps/src/inbd/index.aspx')
    selectType = Select(self.driver.find_element(By.ID,"ctl00_ContentPlaceHolder1_cbo_loaidangkyin"))
    selectType.select_by_value("TB")
    
    selectDest = Select(self.driver.find_element(By.ID,"ctl00_ContentPlaceHolder1_cbo_coso"))
    if selectCampus == "yes":
        selectDest.select_by_value("CS2")

    
    quantityInp = self.driver.find_element(By.ID,"ctl00_ContentPlaceHolder1_txt_soluong")
    quantityInp.send_keys(quantity)

    confirmInp = self.driver.find_element(By.ID,"ctl00_ContentPlaceHolder1_chk_allow")
    if confirm == "yes":
        confirmInp.click()
    
    submitBtn = self.driver.find_element(By.ID,"ctl00_ContentPlaceHolder1_bnt_xacnhan")
    
    if expectedResult == "Success":
        submitBtn.click()
        time.sleep(2)
        cancelPrint = self.driver.find_element(By.NAME,'ctl00$ContentPlaceHolder1$lst_dsphieu$ctl00$bnt_huyphieu')
        cancelPrint.click()
    elif expectedResult == "UnableBtn":        
        assert not submitBtn.is_enabled()
    elif expectedResult == "LackLocation":        
        submitBtn.click()
        time.sleep(2)
        notification = self.driver.find_element(By.CSS_SELECTOR,'.msgBoxContent span')
        assert notification.text == "Chọn nơi nhận kết quả"
    elif expectedResult == "LackQuantity": 
        submitBtn.click()
        time.sleep(2)
        notification = self.driver.find_element(By.CSS_SELECTOR,'.msgBoxContent span')
        assert  "Nhập số lượng" == notification.text               
    elif expectedResult == "InvalidQuantity":        
        submitBtn.click()
        time.sleep(2)
        
        notification = self.driver.find_element(By.CSS_SELECTOR,'.msgBoxContent span')
        assert notification.text == "Số lượng in tối thiểu 1 và tối đa 9"
        
    
  

if __name__ == "__main__":
    excel = FileExcelReader('SecB_printrans_data.xlsx', 'Sheet1')

    test = TestPrintTranscript()
    test.setup_method()
    nRows = excel.getRowCount()
    for row in range(2, nRows + 1):
        selectCampus = excel.readData(row,1)
        quantity = excel.readData(row,2)
        confirm = excel.readData(row,3)
        expectedResult = excel.readData(row,3)

        if quantity is None:
            quantity = ""

            
        try:
            result = test.test_printtranscript(selectCampus,quantity,confirm,expectedResult)
            excel.writeData("Passed",row,5)
        except:
            excel.writeData("Failed",row,5)


    test.teardown_method()